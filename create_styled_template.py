#!/usr/bin/env python3
"""Create a styled students.xlsx.template"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create initial data
data = {
    '#': [1, 2, 3],
    'code': ['STU001', 'STU002', 'STU003'],
    'name': ['Example Student 1', 'Example Student 2', 'Example Student 3'],
    'url': ['', '', '']
}

df = pd.DataFrame(data)

# Save to Excel
df.to_excel('students.xlsx.template', index=False, engine='openpyxl')

# Load the workbook to apply styling
wb = load_workbook('students.xlsx.template')
ws = wb.active

# Define styles
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)
cell_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

# Style header row
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = cell_border

# Style data rows
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
    for cell in row:
        cell.border = cell_border
        cell.alignment = Alignment(vertical='center')

# Set column widths
ws.column_dimensions['A'].width = 6   # #
ws.column_dimensions['B'].width = 12  # code
ws.column_dimensions['C'].width = 30  # name
ws.column_dimensions['D'].width = 15  # url

# Add note about url column
ws['D5'] = "URLs will be generated automatically"
ws['D5'].font = Font(italic=True, color='666666', size=9)

# Freeze the header row
ws.freeze_panes = 'A2'

# Save
wb.save('students.xlsx.template')
print("✓ Created styled students.xlsx.template")
